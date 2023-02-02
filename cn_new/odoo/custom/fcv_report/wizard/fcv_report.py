# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class FcvReport(models.TransientModel):
    _name = "fcv.report"
    _description = "fcv.report"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    fcv_report = fields.Binary('Fcv Report')
    file_name = fields.Char('File Name')
    fcv_printed = fields.Boolean('Fcv Printed')
    warehouse_id = fields.Many2one('stock.warehouse',"Warehouse",required=True)
    
    

    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))
    # Get the product and category filters 

    

    def generate_report(self):

    
        # #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "FCV"
        report_heading = " FCV Report from" + ' ' + datetime.strftime(self.date_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( self.date_end, '%d-%m-%Y')
        
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        # date_range = ws.cell(row=3, column=1, value="Date Range")
        # date_range.alignment = copy(comapny.alignment)
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)


        city = ws.cell(row=3, column=2, value="City")
        city.alignment = copy(comapny.alignment)

        site_name = ws.cell(row=3, column=3, value="Site Name")
        site_name.alignment = Alignment(horizontal='center',vertical='center')

        category_name = ws.cell(row=3, column=4, value="Final Category Name")
        category_name.alignment = Alignment(horizontal='center',vertical='center')

        item_name = ws.cell(row=3, column=5, value="Final Item Name")
        item_name.alignment = Alignment(horizontal='center',vertical='center')

        opening_qty = ws.cell(row=3, column=6, value="Opening Quantity")
        opening_qty.alignment = Alignment(horizontal='center',vertical='center')
        
        opening_value = ws.cell(row=3, column=7, value="Opening Value")
        opening_value.alignment = Alignment(horizontal='center',vertical='center')

        extgr_qty = ws.cell(row=3, column=8, value="External GRN Quantity")
        extgr_qty.alignment = Alignment(horizontal='center',vertical='center')
        
        extgr_val = ws.cell(row=3, column=9, value="External GRN Value")
        extgr_val.alignment = Alignment(horizontal='center',vertical='center')

        dcdone_qty = ws.cell(row=3, column=10, value="Transfer in Quantity")
        dcdone_qty.alignment = Alignment(horizontal='center',vertical='center')
        
        stoin_val = ws.cell(row=3, column=11, value="Transfer in Value")
        stoin_val.alignment = Alignment(horizontal='center',vertical='center')

        stoout_qty = ws.cell(row=3, column=12, value="Transfer out Quantity")
        stoout_qty.alignment = Alignment(horizontal='center',vertical='center')
        
        sto_out_val = ws.cell(row=3, column=13, value="Transfer Out Value")
        sto_out_val.alignment = Alignment(horizontal='center',vertical='center')
        
        intgrn_qty = ws.cell(row=3, column=14, value="GRN Internal Quantity")
        intgrn_qty.alignment = Alignment(horizontal='center',vertical='center')
         
        intgrn_val = ws.cell(row=3, column=15, value="GRN Internal Value")
        intgrn_val.alignment = Alignment(horizontal='center',vertical='center')

        designpo_qty = ws.cell(row=3, column=16, value="Design PO Quantity")
        designpo_qty.alignment = Alignment(horizontal='center',vertical='center')
        
        designpo_val = ws.cell(row=3, column=17, value="Design PO Value")
        designpo_val.alignment = Alignment(horizontal='center',vertical='center')
        
        closing_qty = ws.cell(row=3, column=18, value="Closing Quantity")
        closing_qty.alignment = Alignment(horizontal='center',vertical='center')

        closing_val = ws.cell(row=3, column=19, value="Closing Value")
        closing_val.alignment = Alignment(horizontal='center',vertical='center')

        staffcost_qty = ws.cell(row=3, column=20, value="Staff Cost Quantity")
        staffcost_qty.alignment = Alignment(horizontal='center',vertical='center')

        staffcost_val = ws.cell(row=3, column=21, value="Staff Cost Value")
        staffcost_val.alignment = Alignment(horizontal='center',vertical='center')

        tostore_qty = ws.cell(row=3, column=22, value="To Store to Indian Quantity")
        tostore_qty.alignment = Alignment(horizontal='center',vertical='center')

        tostore_val = ws.cell(row=3, column=23, value="To Store to Indian Value")
        tostore_val.alignment = Alignment(horizontal='center',vertical='center')

        adhocpo_qty = ws.cell(row=3, column=24, value="Adhoc PO Quantity")
        adhocpo_qty.alignment = Alignment(horizontal='center',vertical='center')

        adhocpo_val = ws.cell(row=3, column=25, value="Adhoc PO Value")
        adhocpo_val.alignment = Alignment(horizontal='center',vertical='center')

        grrout_qty = ws.cell(row=3, column=26, value="GRR Out Quantity")
        grrout_qty.alignment = Alignment(horizontal='center',vertical='center')

        grrout_val = ws.cell(row=3, column=27, value="GRR Out Value")
        grrout_val.alignment = Alignment(horizontal='center',vertical='center')

        variance_qty = ws.cell(row=3, column=28, value="Variance Quantity")
        variance_qty.alignment = Alignment(horizontal='center',vertical='center')

        variance_val = ws.cell(row=3, column=29, value="Variance Value")
        variance_val.alignment = Alignment(horizontal='center',vertical='center')

        final_variance = ws.cell(row=3, column=30, value="Final Variance")
        final_variance.alignment = Alignment(horizontal='center',vertical='center')

        togrn_spillover= ws.cell(row=3, column=31, value="TO-GRN spillover")
        togrn_spillover.alignment = Alignment(horizontal='center',vertical='center')

        togrn_spillovers = ws.cell(row=3, column=32, value="TO-GRN spillover")
        togrn_spillovers.alignment = Alignment(horizontal='center',vertical='center')

        stockadj_exp = ws.cell(row=3, column=33, value="Stock Adjustment-Expired")
        stockadj_exp.alignment = Alignment(horizontal='center',vertical='center')

        stockadj_wast = ws.cell(row=3, column=34, value="Stock Adjustment-Wasted")
        stockadj_wast.alignment = Alignment(horizontal='center',vertical='center')

        stockadj_block = ws.cell(row=3, column=35, value="Stock Adjustment-Inventory Added")
        stockadj_block.alignment = Alignment(horizontal='center',vertical='center')


        stockadj_damag = ws.cell(row=3, column=36, value="Stock Adjustment-Damaged")
        stockadj_damag.alignment = Alignment(horizontal='center',vertical='center')


        stockadj_sale = ws.cell(row=3, column=37, value="Stock Adjustment-Sale")
        stockadj_sale.alignment = Alignment(horizontal='center',vertical='center')


        stockadj_closing = ws.cell(row=3, column=38, value="Stock Adjustment-Closing")
        stockadj_closing.alignment = Alignment(horizontal='center',vertical='center')

        variance = ws.cell(row=3, column=39, value="Variance")
        variance.alignment = Alignment(horizontal='center',vertical='center')

        
        # caliculating the opening qty.
        opening_qty_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date < %s and 
                    product_id =%s and
                    warehouse_id = %s  and
                    transaction_types != 'internal'

            group by product_id

            '''
        
        # caliculating the opening value.
        opening_value_query = '''
        select 
                product_id ,
                ABS(sum(product_qty)) as qty,
                ABS(sum(value)) as value
        from 
                inventory_base_report 
        where 
                date < %s and 
                product_id =%s and
                warehouse_id = %s  and
                transaction_types NOT IN  ( 'internal' )

        group by product_id

        '''
        
        # caliculating the grr balance.
        grr_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date > %s  and 
                    date < %s  and 
                    product_id =%s and
                    warehouse_id = %s  and
                    transaction_types = 'p_return' 

            group by product_id



            '''


        # to in 
        # t_receipt

        to_in_query = '''
           select 
                sm.product_id ,sm.product_qty,ib.value
                    
            from 
                    stock_move as sm,
                    stock_picking as sp,
                    inter_company_transfer_ept as ict,
                    inventory_base_report ib

            where   
                    sm.id = ib.move_id and
                    sm.picking_id = sp.id and
                    sp.inter_company_transfer_id = ict.id and
                    sm.date >=%s  and 
                    sm.date <= %s   and 
                    sm.product_id = %s   and 
                    ict.destination_warehouse_id = %s and
                    sp.picking_type_name = 'Delivery Orders' and
                    sp.state = 'done'

            '''

        # To out qty
        to_out_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date > %s  and 
                    date < %s  and 
                    product_id =%s and
                    warehouse_id = %s  and
                    transaction_types = 't_shipment' 

            group by product_id

            '''
        # caliculating the internal grn.
        intgr_query = '''
           select 
                sm.product_id ,sm.product_qty,ib.value
                    
            from 
                    stock_move as sm,
                    stock_picking as sp,
                    inter_company_transfer_ept as ict,
                    inventory_base_report ib

            where   
                    sm.id = ib.move_id and
                    sm.picking_id = sp.id and
                    sp.inter_company_transfer_id = ict.id and
                    sm.date >=%s  and 
                    sm.date <= %s   and 
                    sm.product_id = %s   and 
                    ict.destination_warehouse_id = %s and
                    sp.picking_type_name = 'Receipts' and
                    sp.state = 'done'

            '''

        # Design Po  Quantity
        filtstock = "\'"+self.warehouse_id.code+'/Store'+"\'"
        filtind = "\'"+self.warehouse_id.code+'/Indian'+"\'"


        designpo_query = '''
            select 
                sio.product_uom_qty,sl.name,sl.complete_name,rsu.login,sio.z_map
            from 
                    stock_indent_order_line as sio,
                    stock_indent_order as si,
                    stock_location as sl,
                    res_users as rsu,
                    stock_picking as sp
            where
                    sio.id = si.id and
                    si.name = sp.name and
                    sp.date_done >=%s and
                    sp.date_done <=%s and
                    si.location_id = sl.id and
                    sio.product_id = %s and
                    complete_name in ('''+filtstock+''','''+filtind+''') and
                    si.state = 'done' and
                    sio.create_uid = rsu.id and
                    rsu.login = 'jeevan.s@curefoods.in'




            '''





        # caliculating the staff cost qty.
        staffcost_query = '''
           select 
                sm.product_id ,sm.product_qty,ib.value
                    
            from 
                    stock_move as sm,
                    stock_picking as sp,
                    inter_company_transfer_ept as ict,
                    stock_location as sl,
                    inventory_base_report ib

            where   
                    sm.id = ib.move_id and
                    sm.picking_id = sp.id and
                    sp.location_dest_id =sl.id and 
                    sl.is_staff_location = 't' and
                    sm.date >=%s  and 
                    sm.date <= %s   and 
                    sm.product_id = %s   and 
                    ict.destination_warehouse_id = %s and
                    sp.picking_type_name = 'Internal Transfers' and
                    sp.state = 'done'

            '''

        # caliculating to store to indian qty.
        filtstock = "\'"+self.warehouse_id.code+'/Store'+"\'"
        filtind = "\'"+self.warehouse_id.code+'/Indian'+"\'"
        storeind_query = '''
           select 
                sm.product_id ,sm.product_qty,ib.value,sl.complete_name
                    
            from 
                    stock_move as sm,
                    stock_picking as sp,
                    inter_company_transfer_ept as ict,
                    stock_location as sl,
                    inventory_base_report ib

            where   
                    sm.id = ib.move_id and
                    sm.picking_id = sp.id and
                    sp.location_id = sl.id and
                    complete_name in ('''+filtstock+''','''+filtind+''') and
                    ib.warehouse_id=%s and

                    sm.date >=%s  and 
                    sm.date <= %s   and 
                    sm.product_id = %s   and 
                    ict.destination_warehouse_id = %s and
                    sp.picking_type_name = 'Internal Transfers' and
                    sp.state = 'done'

            '''

        # Design Po  Quantity
        filt = self.warehouse_id.code
        filtstock = "\'"+self.warehouse_id.code+'/Store'+"\'"
        filtind = "\'"+self.warehouse_id.code+'/Indian'+"\'"

        adhocpo_query = '''
            select 
                sio.product_uom_qty,sl.name,sl.complete_name,rsu.login,sio.z_map
            from 
                    stock_indent_order_line as sio,
                    stock_indent_order as si,
                    stock_location as sl,
                    res_users as rsu,
                    stock_picking as sp
            where
                    si.id = sio.id and
                    si.name = sp.name and
                    sp.date_done >=%s and
                    sp.date_done <=%s and
                    si.location_id = sl.id and
                    sio.product_id = %s and
                    sl.complete_name in ('''+filtstock+''','''+filtind+''') and
                    sio.create_uid = rsu.id and
                    rsu.login != 'jeevan.s@curefoods.in' and                


                    si.state = 'done' 
       



            '''



        
        # GR
        # p_receipt
        gr_in_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date > %s  and 
                    date < %s  and 
                    product_id =%s and
                    warehouse_id = %s  and
                    transaction_types = 'p_receipt' 

            group by product_id

            '''

        # To Grn spillover
        grnspill_query = '''
            select 
                    sm.product_qty,sm.z_price
            from 
                    stock_move as sm,
                    stock_picking as sp,
                    stock_picking_type as spt,
                    inter_company_transfer_ept as ict

            where
                    sm.picking_id = sp.ict_id_no and
                    sp.inter_company_transfer_id = ict.id and
                    sp.date_done >=%s  and 
                    sp.date_done <= %s   and 
                    sm.product_id = %s   and
                    sp.picking_type_id = spt.id and
                    spt.warehouse_id = %s and
                    sp.picking_type_name in ('Receipts','Delivery Orders') and
                    sm.date < %s and 
                    sp.state = 'done'                

            '''


        # To Grn spillover
        grnspills_query = '''
            select 
                    sm.product_qty,sm.z_price
            from 
                    stock_move as sm,
                    stock_picking as sp,
                    stock_picking_type as spt,
                    inter_company_transfer_ept as ict

            where
                    sm.picking_id = sp.ict_id_no and
                    sp.inter_company_transfer_id = ict.id and
                    sp.date_done >=%s  and  
                    sm.product_id = %s   and
                    sp.picking_type_id = spt.id and
                    spt.warehouse_id = %s and
                    sp.picking_type_name in ('Receipts','Delivery Orders') and
                    sm.date >= %s and
                    sm.date <= %s and 
                    sp.state = 'done'                

            '''




        # expired quantity 
        expired_query='''
            select 
                    sm.product_id ,sm.product_qty,sm.loss_reason,
                    CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value
            from 
                    stock_move as sm,
                    inventory_base_report ib
            where   
                    sm.id = ib.move_id and
                    sm.date > %s  and 
                    sm.date < %s and
                    sm.product_id = %s and
                    ib.warehouse_id = %s and
                    sm.loss_reason = 'expired' and

                    (sm.location_id =14 or sm.location_dest_id =14)
            '''        


        # wasted quantity 
        wasted_query='''
            select 
                    sm.product_id ,sm.product_qty,sm.loss_reason,
                    CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value
            from 
                    stock_move as sm,
                    inventory_base_report ib
            where   
                    sm.id = ib.move_id and
                    sm.date > %s  and 
                    sm.date < %s and
                    sm.product_id = %s and
                    ib.warehouse_id = %s and
                    sm.loss_reason = 'wasted' and

                    (sm.location_id =14 or sm.location_dest_id =14)
            '''
    
        # inventory added quantity 
        invad_query='''
            select 
                    sm.product_id ,sm.product_qty,sm.loss_reason,
                    CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value
            from 
                    stock_move as sm,
                    inventory_base_report ib
            where   
                    sm.id = ib.move_id and
                    sm.date > %s  and 
                    sm.date < %s and
                    sm.product_id = %s and
                    ib.warehouse_id = %s and
                    sm.loss_reason = 'inventory_added' and

                    (sm.location_id =14 or sm.location_dest_id =14)
            '''


        # damaged quantity 
        damaged_query='''
            select 
                    sm.product_id ,sm.product_qty,sm.loss_reason,
                    CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value

            from 
                    stock_move as sm,
                    inventory_base_report ib
            where   
                    sm.id = ib.move_id and
                    sm.date > %s  and 
                    sm.date < %s and
                    sm.product_id = %s and
                    ib.warehouse_id = %s and
                    sm.loss_reason = 'damaged' and
                    (sm.location_id =14 or sm.location_dest_id =14)
            '''

        # sale quantity 
        sale_query='''
            select 
                    sm.product_id ,sm.product_qty,sm.loss_reason,
                    CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value

            from 
                    stock_move as sm,
                    inventory_base_report ib
            where   
                    sm.id = ib.move_id and
                    sm.date > %s  and 
                    sm.date < %s and
                    sm.product_id = %s and
                    ib.warehouse_id = %s and
                    sm.loss_reason = 'sale' and
                    (sm.location_id =14 or sm.location_dest_id =14)
            '''

        # closing quantity 
        closing_query='''
            select 
                    sm.product_id ,sm.product_qty,sm.loss_reason,
                    CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value

            from 
                    stock_move as sm,
                    inventory_base_report ib
            where   
                    sm.id = ib.move_id and
                    sm.date > %s  and 
                    sm.date < %s and
                    sm.product_id = %s and
                    ib.warehouse_id = %s and
                    sm.loss_reason = 'closing' and
                    (sm.location_id =14 or sm.location_dest_id =14)
            '''        

        # caliculating the closing qty.
        closing_qty_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date <= %s and 
                    product_id =%s and
                    warehouse_id = %s  and
                    transaction_types != 'internal'

            group by product_id

            '''


        # caliculating the closing value.
        closing_value_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date <= %s and 
                    product_id =%s and
                    warehouse_id = %s  and
                    transaction_types NOT IN  ( 'internal' )

            group by product_id

            '''

        # Variance quantity 
        variance_query='''
            select 
                    sm.product_id ,sm.product_qty,sm.loss_reason,\
                    CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value

            from 
                    stock_move as sm,
                    inventory_base_report ib
            where   
                    sm.id = ib.move_id and
                    sm.date > %s  and 
                    sm.date < %s and
                    sm.product_id = %s and
                    ib.warehouse_id = %s and
                    ib.transaction_types IN ('positive','negative') and 

                    (sm.location_id =14 or sm.location_dest_id =14)
            '''


        
        row_c=4
        sl_num=1
        
        for each_base in self.env['product.product'].search([]):

            # Get the product name,uom and the product Categery
            product_quer='''
            select 
                    pt.name as p_name,
                    uu.name as uu_name,
                    pc.complete_name  as pc_name
            from  
                product_product as pp,
                product_template as pt,
                uom_uom as uu,
                product_category as pc
            where 
                product_tmpl_id = pt.id and 
                pp.id =%s and
                uu.id = pt.uom_id and
                pc.id = pt.categ_id
                '''        
            self.env.cr.execute(product_quer, [each_base.id,])
            product = self.env.cr.dictfetchall() 

            # opening qty
            opening_qt_params = (self.date_start,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(opening_qty_query,opening_qt_params)
            opening_qt_result = self.env.cr.dictfetchall()

            # opening Value
            opening_value_params = (self.date_start,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(opening_value_query,opening_value_params)
            opening_value_result = self.env.cr.dictfetchall()
            
            # Grr balance
            grr_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(grr_query,grr_params)
            grr_result = self.env.cr.dictfetchall()

            # to in balance
            to_in_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(to_in_query,to_in_params)
            to_in_result = self.env.cr.dictfetchall()

            # To Out balance
            to_out_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(to_out_query,to_out_params)
            to_out_result = self.env.cr.dictfetchall()

            # Grin balance
            gr_in_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(gr_in_query,gr_in_params)
            gr_in_result = self.env.cr.dictfetchall()


            # To Grn spillover
            grnsill_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id,self.date_start)
            self.env.cr.execute(grnspill_query,grnsill_params)
            grnspill_result = self.env.cr.dictfetchall()


            # To Grn spillovers
            grnsills_params = (self.date_end,each_base.id,self.warehouse_id.id,self.date_start,self.date_end)
            self.env.cr.execute(grnspills_query,grnsills_params)
            grnspills_result = self.env.cr.dictfetchall()            

            # Internal Grn
            intgr_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(intgr_query,intgr_params)
            intgr_result = self.env.cr.dictfetchall()  

            # Design Po
            designpo_params = (self.date_start,self.date_end,each_base.id)
            self.env.cr.execute(designpo_query,designpo_params)
            designpo_result = self.env.cr.dictfetchall()              


            # Staff Cost 
            staffcost_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(staffcost_query,staffcost_params)
            staffcost_result = self.env.cr.dictfetchall()  


            # To store to indian 
            storeind_params = (self.warehouse_id.id,self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(storeind_query,storeind_params)
            storeind_result = self.env.cr.dictfetchall()  
                      
            # Adhoc Po
            adhocpo_params = (self.date_start,self.date_end,each_base.id)
            self.env.cr.execute(adhocpo_query,adhocpo_params)
            adhocpo_result = self.env.cr.dictfetchall() 

           # expired balance
            expired_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(expired_query,expired_params)
            expired_result = self.env.cr.dictfetchall()


           # wasted balance
            wasted_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(wasted_query,wasted_params)
            wasted_result = self.env.cr.dictfetchall()


           # inventory added balance
            invad_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(invad_query,invad_params)
            invad_result = self.env.cr.dictfetchall()           

            # damaged balance
            damaged_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(damaged_query,damaged_params)
            damaged_result = self.env.cr.dictfetchall()

            # sale balance
            sale_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(sale_query,sale_params)
            sale_result = self.env.cr.dictfetchall()

            # closing balance
            closing_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(closing_query,closing_params)
            closing_result = self.env.cr.dictfetchall()

            expired_qty = 0.0
            expired_value = 0.0
            if expired_result:
                for each_expired in expired_result:
                    expired_value += -(each_expired['value']) if each_expired['value'] <= 0 else each_expired['value']
                    expired_qty += each_expired['product_qty']


            
            wasted_qty = 0.0
            wasted_value = 0.0
            if wasted_result:
                for each_waste in wasted_result:
                    wasted_value += -(each_waste['value']) if each_waste['value'] <= 0 else each_waste['value']
                    wasted_qty += each_waste['product_qty']

            invad_qty = 0.0
            invad_value = 0.0
            if invad_result:
                for each_invad in invad_result:
                    invad_value += -(each_invad['value']) if each_invad['value'] <= 0 else each_invad['value']
                    invad_qty += each_invad['product_qty']

            d_qty = 0.0
            d_value = 0.0
            if damaged_result:
                for each_damaged in damaged_result:
                    d_value += -(each_damaged['value']) if each_damaged['value'] <= 0 else each_damaged['value']
                    d_qty += each_damaged['product_qty']

            sale_qty = 0.0
            sale_value = 0.0
            if sale_result:
                for each_sale in sale_result:
                    sale_value += -(each_sale['value']) if each_sale['value'] <= 0 else each_sale['value']
                    sale_qty += each_sale['product_qty']

            closing_qty = 0.0
            closing_value = 0.0
            if closing_result:
                for each_closing in closing_result:
                    closing_value += -(each_closing['value']) if each_closing['value'] <= 0 else each_closing['value']
                    closing_qty += each_closing['product_qty']




            # Closing qty
            closing_qty_params = (self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(closing_qty_query,closing_qty_params)
            closing_qty_result = self.env.cr.dictfetchall()

            # Closing value
            cclosing_value_params = (self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(closing_value_query,cclosing_value_params)
            closing_value_result = self.env.cr.dictfetchall()

            # Variance balance
            variance_params = (self.date_start,self.date_end,each_base.id,self.warehouse_id.id)
            self.env.cr.execute(variance_query,variance_params)
            variance_result = self.env.cr.dictfetchall()            



            t_in_qty = 0.0
            t_in_value = 0.0
            if to_in_result:
                for each_to_in in to_in_result:
                    t_in_value += each_to_in['value']
                    t_in_qty += each_to_in['product_qty']
            

            grspill_qty = 0.0
            grspill_value = 0.0
            if grnspill_result:
                for each_grspill in grnspill_result:
                    grspill_value += each_grspill['product_qty'] * each_grspill['z_price']
                    grspill_qty += each_grspill['product_qty']

            grspills_qty = 0.0
            grspills_value = 0.0
            if grnspill_result:
                for each_grspills in grnspills_result:
                    grspills_value += each_grspills['product_qty'] * each_grspills['z_price']
                    grspills_qty += each_grspills['product_qty']



            int_gr_qty = 0.0
            int_gr_value = 0.0
            if intgr_result:
                for each_gr_in in intgr_result:
                    int_gr_value += each_gr_in['value']
                    int_gr_qty += each_gr_in['product_qty']

            design_po_qty = 0.0
            design_po_value = 0.0
            if designpo_result:
                for each_designpo_in in designpo_result:
                    design_po_value += each_designpo_in['product_uom_qty']*(each_designpo_in['z_map'] if each_designpo_in['z_map'] else 0.0)
                    design_po_qty += each_designpo_in['product_uom_qty']


            staff_cost_qty = 0.0
            staff_cost_value = 0.0
            if staffcost_result:
                for each_staffcost in staffcost_result:
                    staff_cost_value += each_staffcost['value']
                    staff_cost_qty += each_staffcost['product_qty']


            store_ind_qty = 0.0
            store_ind_value = 0.0
            if storeind_result:
                for each_storeind in storeind_result:
                    store_ind_value += each_storeind['value']
                    store_ind_qty += each_storeind['product_qty']

            adhoc_po_qty = 0.0
            adhoc_po_value = 0.0
            if adhocpo_result:
                
                for each_adhocpo in adhocpo_result:
                    adhoc_po_value += each_adhocpo['product_uom_qty']*(each_adhocpo['z_map'] if each_adhocpo['z_map'] else 0.0)
                    adhoc_po_qty += each_adhocpo['product_uom_qty']

            v_qty = 0.0
            v_value = 0.0
            if variance_result:
                for each_variance in variance_result:
                    v_value += each_variance['value'] 
                    v_qty += each_variance['product_qty']


                                           

            p_op =opening_qt_result[0]['qty'] if opening_qt_result  else 0.0

            p_to_out = -(to_out_result[0]['qty']) if to_out_result  else 0.0 

            p_grr = grr_result[0]['qty'] if grr_result  else 0.0

            to_storeval = storeind_result[0]['value']  if storeind_result  else 0.0 

            p_clo=closing_qty_result[0]['qty'] if closing_qty_result  else 0.0

            p_out_val=(to_out_result[0]['value'])   if to_out_result  else 0.0

            staff_costval = staffcost_result[0]['value']  if staffcost_result  else 0.0 

            p_gr =gr_in_result[0]['qty'] if gr_in_result  else 0.0

            vari_qty = (p_op + p_gr + t_in_qty) - (p_to_out + p_grr + p_clo) - (design_po_qty + staff_cost_qty)

            

            vari_val = (opening_value_result[0]['value'] if opening_value_result  else 0.0 + gr_in_result[0]['value']  if gr_in_result  else 0.0 + t_in_value) - (-(p_out_val) if p_out_val < 0 else p_out_val + grr_result[0]['value']  if grr_result  else 0.0 + closing_value_result[0]['value']  if closing_value_result  else 0.0) - (design_po_value + staffcost_result[0]['value']  if staffcost_result  else 0.0 )

            varia_value =vari_val + grspill_value - grspills_value

            sl_val = ws.cell(row=row_c, column=1, value=sl_num)
            sl_val.alignment = Alignment(horizontal='center',vertical='center')

            
            city = ws.cell(row=row_c, column=2, value=self.warehouse_id.city_id.name)


            site_name = ws.cell(row=row_c, column=3, value=self.warehouse_id.name)

            category_name = ws.cell(row=row_c, column=4, value=product[0]['pc_name'])

            item_name = ws.cell(row=row_c, column=5, value=product[0]['p_name'])

            opening_qty = ws.cell(row=row_c, column=6, value=p_op)

            opening_value = ws.cell(row=row_c, column=7, value=opening_value_result[0]['value'] if opening_value_result  else 0.0)


            # print("gr_in_result[0]['qty']gr_in_result[0]['qty']gr_in_result[0]['qty']",gr_in_result)
            extgr_qty = ws.cell(row=row_c, column=8, value=p_gr)
            extgr_val = ws.cell(row=row_c, column=9, value=gr_in_result[0]['value']  if gr_in_result  else 0.0)

            dcdone_qty = ws.cell(row=row_c, column=10, value=t_in_qty)
            stoin_val = ws.cell(row=row_c, column=11, value=-(t_in_value) if t_in_value < 0 else t_in_value )

            stoout_qty = ws.cell(row=row_c, column=12, value=-(p_to_out) if p_to_out < 0 else p_to_out )
            # pdb.set_trace()
            sto_out_val = ws.cell(row=row_c, column=13, value= -(p_out_val) if p_out_val < 0 else p_out_val)

            # pdb.set_trace()


            intgrn_qty = ws.cell(row=row_c, column=14, value=int_gr_qty)
            intgrn_val = ws.cell(row=row_c, column=15, value=-(int_gr_value) if int_gr_value < 0 else int_gr_value )

            designpo_qty = ws.cell(row=row_c, column=16, value=design_po_qty)
            designpo_val = ws.cell(row=row_c, column=17, value=design_po_value)

            closing_qty = ws.cell(row=row_c, column=18, value=p_clo )
            closing_val = ws.cell(row=row_c, column=19, value=closing_value_result[0]['value']  if closing_value_result  else 0.0 )

            staffcost_qty = ws.cell(row=row_c, column=20, value=staff_cost_qty)
            staffcost_val = ws.cell(row=row_c, column=21, value= -(staff_costval) if staff_costval < 0 else staff_costval)

            tostore_qty = ws.cell(row=row_c, column=22, value=store_ind_qty)
            tostore_val = ws.cell(row=row_c, column=23, value= -(to_storeval) if to_storeval <0 else to_storeval)


            adhocpo_qty = ws.cell(row=row_c, column=24, value=adhoc_po_qty)
            adhocpo_val = ws.cell(row=row_c, column=25, value=adhoc_po_value)

            grrout_qty = ws.cell(row=row_c, column=26, value=p_grr)
            grrout_val = ws.cell(row=row_c, column=27, value=grr_result[0]['value']  if grr_result  else 0.0)

            variance_qty = ws.cell(row=row_c, column=28, value= -(vari_qty) if vari_qty < 0 else vari_qty)
            variance_val = ws.cell(row=row_c, column=29, value= -(vari_val) if vari_val < 0 else vari_val) 

            final_variance = ws.cell(row=row_c, column=30, value= -(varia_value) if varia_value<0 else varia_value) 
            togrn_spillover = ws.cell(row=row_c, column=31, value= grspill_value)

            togrn_spillovers = ws.cell(row=row_c, column=32, value=grspills_value)
            stockadj_exp = ws.cell(row=row_c, column=33, value=-(expired_value) if expired_value <0 else expired_value )

            stockadj_wast = ws.cell(row=row_c, column=34, value=-(wasted_value) if wasted_value <0 else wasted_value )
            stockadj_block = ws.cell(row=row_c, column=35, value=-(invad_value) if invad_value <0 else invad_value )

            stockadj_damag = ws.cell(row=row_c, column=36, value=-(d_value)  if d_value < 0 else d_value )

            stockadj_sale = ws.cell(row=row_c, column=37, value=-(sale_value)  if sale_value < 0 else sale_value )

            stockadj_closing = ws.cell(row=row_c, column=38, value=-(closing_value)  if closing_value < 0 else closing_value )


            variance = ws.cell(row=row_c, column=39, value=v_value)


            # sl_num +=1
            row_c +=1
            sl_num +=1
    
        


        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.fcv_report = excel_file
        self.fcv_printed = True
        self.file_name = "FCV Report.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'fcv.report',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
